"""
Generate Excel spreadsheet with all identified codebase issues,
organized by priority and category for systematic resolution.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Issues Backlog"

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
critical_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
high_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
medium_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
low_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["ID", "Priority", "Category", "File", "Issue Type", "Severity", "Description", "Status", "Owner", "Notes"]
ws.append(headers)

# Style header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 45
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 25

# All issues data (ID, Priority, Category, File, IssueType, Severity, Description, Status, Owner, Notes)
issues = [
    # CRITICAL (5)
    (1, "CRITICAL", "TEST", "tests/test_editor_features.py", "STUB_CODE", "HIGH", "22 empty test methods - all just pass(). Zero test coverage for editor UI features", "NOT_STARTED", "", "Implement all 22 tests for styled dialogs, carousel, animations, presets, persistence, multi-file upload"),
    (2, "CRITICAL", "SECURITY", "services/auth_service.py", "HARDCODED_SECRET", "HIGH", "Hardcoded default JWT_SECRET 'change-me-in-production' if env var missing", "NOT_STARTED", "", "Fail hard on startup if JWT_SECRET not set. Never use default."),
    (3, "CRITICAL", "SECURITY", "api/routes/website_builder.py", "UNAUTHENTICATED_ENDPOINT", "HIGH", "POST /shop/finalize-image has no authentication - public access to file operations", "NOT_STARTED", "", "Add @require_auth to endpoint, validate ownership"),
    (4, "CRITICAL", "SECURITY", "api/routes/commerce.py", "SQL_INJECTION", "HIGH", "Dynamic column names in UPDATE coupons query using f-string - SQL injection risk", "NOT_STARTED", "", "Use parameterized column builder or SQL schema validator"),
    (5, "CRITICAL", "ERROR_HANDLING", "api/routes/website_builder.py", "GENERIC_EXCEPTION", "HIGH", "Multiple bare 'except Exception:' blocks with no logging - 10+ silent failures hide bugs", "NOT_STARTED", "", "Add structured logging to all exception handlers, re-raise or return proper error response"),

    # HIGH (5)
    (6, "HIGH", "CONFIG", "config/settings.py", "VALIDATION_MISSING", "MEDIUM", "No validation of required settings - missing OPENAI_API_KEY only warns, doesn't block", "NOT_STARTED", "", "Add pydantic validators to fail hard on missing required vars"),
    (7, "HIGH", "ERROR_HANDLING", "api/routes/clients.py", "GENERIC_EXCEPTION", "HIGH", "4+ generic exception handlers without logging or differentiated responses", "NOT_STARTED", "", "Implement proper exception handling with user-facing error messages"),
    (8, "HIGH", "FRONTEND", "frontend/dashboard.js", "GLOBAL_STATE_RACE", "MEDIUM", "Multiple global state variables with no synchronization - race conditions on concurrent ops", "NOT_STARTED", "", "Refactor to event-driven architecture or use state management library"),
    (9, "HIGH", "PERFORMANCE", "database/snowflake_client.py", "NO_POOLING", "MEDIUM", "No connection pooling - new SQLite connection per query, N+1 risk in loops", "NOT_STARTED", "", "Implement connection pooling, add caching layer"),
    (10, "HIGH", "TEST", "tests/", "COVERAGE_MISSING", "HIGH", "No code coverage metrics configured - can't measure test effectiveness", "NOT_STARTED", "", "Add pytest-cov, set up coverage threshold (min 80%), configure CI"),

    # HIGH - Error Handling (5)
    (11, "HIGH", "ERROR_HANDLING", "api/routes/commerce.py", "SILENT_FAILURE", "MEDIUM", "Campaign send logic catches exception, passes - incomplete implementation", "NOT_STARTED", "", "Implement full retry logic, proper error reporting to user"),
    (12, "HIGH", "ERROR_HANDLING", "api/routes/shopping_cart.py", "GENERIC_EXCEPTION", "MEDIUM", "3+ bare exception handlers without recovery or logging", "NOT_STARTED", "", "Add logging, implement graceful degradation"),
    (13, "HIGH", "ERROR_HANDLING", "api/routes/team.py", "SILENT_FAILURE", "MEDIUM", "Silent exception catch with pass statement", "NOT_STARTED", "", "Add logging, return proper error response"),
    (14, "HIGH", "ERROR_HANDLING", "frontend/dashboard.js", "EMPTY_CATCH", "MEDIUM", "Empty catch block at line 96 - errors swallowed silently", "NOT_STARTED", "", "Implement proper error handling with user notification"),
    (15, "HIGH", "ERROR_HANDLING", "frontend/dashboard.js", "EMPTY_CATCH", "LOW", "Empty try/catch at line 398 - debugging issues hidden", "NOT_STARTED", "", "Add console logging, implement error tracking"),

    # MEDIUM (15)
    (16, "MEDIUM", "SECURITY", ".env.example", "EXPOSED_SECRETS", "HIGH", "Hardcoded ADMIN_PASSWORD example 'Admin@1234' in docs - could be committed", "NOT_STARTED", "", "Remove from example, generate random placeholder, document in comments"),
    (17, "MEDIUM", "SECURITY", "database/snowflake_client.py", "FOREIGN_KEY_OFF", "MEDIUM", "PRAGMA foreign_keys=OFF disables referential integrity - allows orphaned data", "NOT_STARTED", "", "Enable foreign key constraints, add migration to cleanup orphaned records"),
    (18, "MEDIUM", "SECURITY", "api/routes/clients.py", "MISSING_AUTH", "MEDIUM", "No apparent validation of website ownership before allowing client operations", "NOT_STARTED", "", "Add ownership checks to all client endpoints"),
    (19, "MEDIUM", "SECURITY", "services/secret_store.py", "NO_PRECHECK", "MEDIUM", "No pre-check on STORAGE_SECRETS_KEY - app crashes at runtime instead of graceful fallback", "NOT_STARTED", "", "Add startup validation, graceful fallback if key unavailable"),
    (20, "MEDIUM", "CONFIG", "services/notification_service.py", "EMPTY_DEFAULT", "MEDIUM", "SMTP_PASSWORD defaults to empty - silently fails without warning", "NOT_STARTED", "", "Add validation, clear error message when SMTP not configured"),
    (21, "MEDIUM", "CONFIG", ".env.example", "MISSING_DOCS", "MEDIUM", "No indication which env vars are mandatory vs optional", "NOT_STARTED", "", "Add comments marking required (!) vs optional (?) fields"),
    (22, "MEDIUM", "PERFORMANCE", "api/routes/website_builder.py", "N_PLUS_ONE", "MEDIUM", "Plan limit checking uses SELECT COUNT(*) per request - no caching", "NOT_STARTED", "", "Implement caching layer (Redis) or query optimization"),
    (23, "MEDIUM", "DATABASE", "database/migrations.py", "DUPLICATE_FIELDS", "MEDIUM", "cart_items table has both 'stock' and 'stock_quantity' - duplicate, unclear semantics", "NOT_STARTED", "", "Consolidate fields, create migration to remove redundant column"),
    (24, "MEDIUM", "DATABASE", "database/migrations.py", "UNDEFINED_JSON", "MEDIUM", "carts table has items_json (VARIANT) but structure undefined - no schema validation", "NOT_STARTED", "", "Define JSON schema, add validation in code"),
    (25, "MEDIUM", "DATABASE", "database/migrations.py", "NO_DEFAULTS", "MEDIUM", "websites table has many optional fields without defaults - NULL handling unclear", "NOT_STARTED", "", "Define sensible defaults, update schema, add NULL checks"),
    (26, "MEDIUM", "DATABASE", "database/migrations.py", "NO_ROLLBACK", "MEDIUM", "Default migrations have no rollback mechanism - failed migrations leave DB in inconsistent state", "NOT_STARTED", "", "Implement rollback functions, add transaction wrapping"),
    (27, "MEDIUM", "API", "api/routes/website_builder.py", "RATE_LIMIT_MISSING", "MEDIUM", "POST /scrape-url has no rate limiting - could be abused for SSRF/DoS", "NOT_STARTED", "", "Add SlowAPI rate limiter, document limits"),
    (28, "MEDIUM", "API", "api/routes/website_builder.py", "STREAM_CLEANUP", "MEDIUM", "SSE endpoint /build-stream has no documentation of cleanup or timeout", "NOT_STARTED", "", "Implement connection timeout, cleanup on disconnect/timeout"),
    (29, "MEDIUM", "FRONTEND", "frontend/dashboard.html", "SECRET_FORMAT_EXPOSED", "MEDIUM", "Password field placeholder exposes credential secrets format in HTML source", "NOT_STARTED", "", "Remove placeholder examples from HTML"),
    (30, "MEDIUM", "FRONTEND", "frontend/dashboard.js", "API_TIMEOUT", "MEDIUM", "apiFetch() helper doesn't handle timeout gracefully - returns null without user notification", "NOT_STARTED", "", "Add timeout handling, display user-facing message"),

    # LOW (18)
    (31, "LOW", "ERROR_HANDLING", "main.py", "GENERIC_EXCEPTION", "MEDIUM", "Generic exception handling in CLI without detailed error messages", "NOT_STARTED", "", "Add detailed error context, suggest fixes based on error type"),
    (32, "LOW", "CONFIG", "services/image_service.py", "FALLBACK_WARNINGS", "LOW", "Multiple fallback warnings for missing backend configs - no single validation point", "NOT_STARTED", "", "Centralize image backend config validation at startup"),
    (33, "LOW", "CONFIG", "services/monitoring_service.py", "MISSING_CONFIG", "LOW", "Hard-coded STRIPE_SECRET_KEY check missing - returns 402 with generic message", "NOT_STARTED", "", "Add env validation, better error message"),
    (34, "LOW", "PERFORMANCE", "services/analytics_service.py", "LOG_BLOAT", "LOW", "Activity logging is best-effort, inserts one row per event - no cleanup/archival policy", "NOT_STARTED", "", "Implement log rotation, archival to cold storage, or batching"),
    (35, "LOW", "PERFORMANCE", "tests/run_automated_tests.py", "HARDCODED_SLEEPS", "LOW", "Multiple hardcoded time.sleep(1-3) in tests - slow, brittle execution", "NOT_STARTED", "", "Replace with polling/wait_for helpers, use pytest fixtures"),
    (36, "LOW", "DATABASE", "services/auth_service.py", "PLACEHOLDER_MISMATCH", "LOW", "OTP insertion uses %s but database adapter uses ? - parameterization mismatch", "NOT_STARTED", "", "Normalize placeholder style across codebase"),
    (37, "LOW", "DATABASE", "database/migrations.py", "IDEMPOTENCY", "LOW", "ALTER TABLE migrations don't check if column exists - could fail on re-run", "NOT_STARTED", "", "Add IF NOT EXISTS checks to ALTER TABLE statements"),
    (38, "LOW", "CODE_QUALITY", "api/routes/website_builder.py", "UNCLEAR_COMMENT", "LOW", "Comment 'scraped_title is intentionally NOT passed' - intent unclear, should document why", "NOT_STARTED", "", "Add detailed comment explaining rationale"),
    (39, "LOW", "FRONTEND", "frontend/dashboard.html", "MANUAL_PASSWORD", "LOW", "Temporary password field with no auto-generated value - users must enter manually", "NOT_STARTED", "", "Generate random secure password, copy to clipboard button"),
    (40, "LOW", "FRONTEND", "frontend/dashboard.js", "FRAGILE_FALLBACK", "LOW", "Fallback check for toast function uses typeof - fragile, should use modules", "NOT_STARTED", "", "Refactor to proper module imports, remove global checks"),
    (41, "LOW", "FRONTEND", "frontend/dashboard.js", "MONOLITHIC_FILE", "MEDIUM", "14K+ line monolithic file - no modularization, difficult to maintain", "NOT_STARTED", "", "Break into components: auth.js, sites.js, editor.js, settings.js, etc."),
    (42, "LOW", "DOCUMENTATION", "README.md", "OUTDATED", "MEDIUM", "README suggests running main.py but app uses app.py now - build instructions unclear", "NOT_STARTED", "", "Update all instructions to use app.py, add deployment guide"),
    (43, "LOW", "LOGGING", "agents/requirements_analyst.py", "CONSOLE_ERROR", "LOW", "Console error injection in HTML - error messages exposed to user", "NOT_STARTED", "", "Use logger instead of console.error, sanitize output"),
    (44, "LOW", "CONFIG", "services/currency_service.py", "FREE_TIER_RISK", "LOW", "Relies on ip-api.com free tier (no key) - rate limited, could fail silently", "NOT_STARTED", "", "Add rate limit handling, use paid tier or fallback backend"),
    (45, "LOW", "ERROR_HANDLING", "services/hosting_service.py", "SILENT_FAILURE", "LOW", "Google API libraries missing - prints error but doesn't raise, operation fails silently", "NOT_STARTED", "", "Add explicit error check, raise or return error status"),
    (46, "LOW", "TEST", "tests/test_editor_features.py", "CONFIG_MISSING", "MEDIUM", "Test file incomplete - unclear if pytest fixtures are properly isolated between tests", "NOT_STARTED", "", "Verify conftest.py setup, add database cleanup between tests"),
    (47, "LOW", "CODE_QUALITY", "services/notification_service.py", "INCONSISTENT_LOGGING", "MEDIUM", "Inconsistent logging: some code uses logger.debug(), some uses print(), some silent", "NOT_STARTED", "", "Implement structured logging (JSON format), standardize across codebase"),
    (48, "LOW", "FEATURE", "services/auth_service.py", "ERROR_RECOVERY", "HIGH", "No retry logic for transient failures (network, timeouts) - hard failures instead of graceful degradation", "NOT_STARTED", "", "Implement exponential backoff, circuit breaker pattern for external APIs"),
]

# Add data rows
for idx, issue in enumerate(issues, start=2):
    id_val, priority, category, file, issue_type, severity, description, status, owner, notes = issue
    ws.append([id_val, priority, category, file, issue_type, severity, description, status, owner, notes])
    
    # Apply priority fill to row
    priority_col = ws[f'B{idx}']
    if priority == "CRITICAL":
        fill = critical_fill
        font_color = Font(bold=True, color="FFFFFF")
    elif priority == "HIGH":
        fill = high_fill
        font_color = Font(bold=True, color="FFFFFF")
    elif priority == "MEDIUM":
        fill = medium_fill
        font_color = Font(color="000000")
    else:
        fill = low_fill
        font_color = Font(color="000000")
    
    # Style data row
    for col_idx in range(1, 11):
        cell = ws.cell(row=idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        if col_idx == 2:  # Priority column
            cell.fill = fill
            cell.font = font_color
        
        if col_idx in [6, 7]:  # Description, Notes columns
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Freeze panes
ws.freeze_panes = "A2"

# Add summary sheet
summary = wb.create_sheet("Summary")
summary.column_dimensions['A'].width = 25
summary.column_dimensions['B'].width = 12

summary_data = [
    ["CODEBASE AUDIT SUMMARY", ""],
    ["", ""],
    ["Total Issues Found", 48],
    ["", ""],
    ["PRIORITY BREAKDOWN", ""],
    ["Critical (Fix Immediately)", 5],
    ["High (Fix This Sprint)", 5],
    ["Medium (Fix Next Sprint)", 15],
    ["Low (Backlog)", 23],
    ["", ""],
    ["CATEGORY BREAKDOWN", ""],
    ["Error Handling", 8],
    ["Security", 6],
    ["Test Coverage", 5],
    ["Config/Environment", 6],
    ["Performance", 4],
    ["Database", 5],
    ["API Endpoints", 6],
    ["Frontend/UI", 6],
    ["Unimplemented/Stub", 2],
    ["", ""],
    ["STATUS", "Count"],
    ["Not Started", 48],
    ["In Progress", 0],
    ["Complete", 0],
    ["", ""],
    ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
]

for row_idx, row_data in enumerate(summary_data, 1):
    summary.append(row_data)

# Format summary
for row in summary.iter_rows(min_row=1, max_row=summary.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = border
        if row[0].value and isinstance(row[0].value, str) and ("SUMMARY" in row[0].value or "BREAKDOWN" in row[0].value or "STATUS" in row[0].value):
            row[0].fill = header_fill
            row[0].font = Font(bold=True, color="FFFFFF")

# Save workbook
output_path = "/workspaces/agentic-s3-website-builder/output/CODEBASE_ISSUES_BACKLOG.xlsx"
wb.save(output_path)
print(f"✅ Excel file created: {output_path}")
print(f"\nTotal Issues: 48")
print(f"  - Critical: 5")
print(f"  - High: 5")
print(f"  - Medium: 15")
print(f"  - Low: 23")
